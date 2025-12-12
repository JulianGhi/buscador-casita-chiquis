#!/usr/bin/env python3
"""
Completa datos faltantes en el Excel scrapeando los links de cada propiedad.
Soporta: MercadoLibre, Argenprop, Zonaprop
"""

import asyncio
import re
import time
import httpx
from bs4 import BeautifulSoup
import openpyxl
from playwright.async_api import async_playwright

EXCEL_PATH = 'data/seguimiento_propiedades_v3.xlsx'


def scrape_argenprop(url):
    """Scrapea Argenprop (SSR, sin protección)"""
    try:
        resp = httpx.get(url, follow_redirects=True,
                        headers={'User-Agent': 'Mozilla/5.0'}, timeout=10)
        if resp.status_code != 200:
            return {'error': f'Status {resp.status_code}'}

        soup = BeautifulSoup(resp.text, 'lxml')
        data = {}

        precio = soup.select_one('.titlebar__price')
        if precio:
            txt = precio.text.strip()
            match = re.search(r'[\d.]+', txt.replace('.', ''))
            if match:
                data['precio'] = int(match.group())

        ubicacion = soup.select_one('.titlebar__address')
        if ubicacion:
            data['direccion'] = ubicacion.text.strip()

        for li in soup.select('.property-features li'):
            txt = li.text.strip()
            if 'm² cub' in txt.lower():
                match = re.search(r'(\d+)', txt)
                if match:
                    data['m2_cub'] = int(match.group(1))
            elif 'm² tot' in txt.lower():
                match = re.search(r'(\d+)', txt)
                if match:
                    data['m2_tot'] = int(match.group(1))

        return data
    except Exception as e:
        return {'error': str(e)}


def scrape_mercadolibre(url):
    """Scrapea MercadoLibre (httpx directo)"""
    try:
        resp = httpx.get(url, follow_redirects=True,
                        headers={'User-Agent': 'Mozilla/5.0'}, timeout=10)
        if resp.status_code != 200:
            return {'error': f'Status {resp.status_code}'}

        soup = BeautifulSoup(resp.text, 'lxml')
        data = {}

        precio = soup.select_one('.andes-money-amount__fraction')
        if precio:
            data['precio'] = int(precio.text.strip().replace('.', ''))

        for row in soup.select('tr.andes-table__row'):
            header = row.select_one('th')
            value = row.select_one('td')
            if header and value:
                h = header.text.strip().lower()
                v = value.text.strip()
                if 'superficie cubierta' in h:
                    match = re.search(r'(\d+)', v)
                    if match:
                        data['m2_cub'] = int(match.group(1))
                elif 'superficie total' in h:
                    match = re.search(r'(\d+)', v)
                    if match:
                        data['m2_tot'] = int(match.group(1))
                elif 'ambientes' in h:
                    match = re.search(r'(\d+)', v)
                    if match:
                        data['amb'] = int(match.group(1))

        ubicacion = soup.select_one('.ui-vip-location a')
        if ubicacion:
            data['barrio'] = ubicacion.text.strip()

        return data
    except Exception as e:
        return {'error': str(e)}


async def scrape_zonaprop(url):
    """Scrapea Zonaprop (requiere Playwright, Cloudflare)"""
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False)
            page = await browser.new_page()

            await page.goto(url, wait_until='domcontentloaded')
            await page.wait_for_timeout(10000)

            data = {}

            precio_el = await page.query_selector('.price-items span')
            if precio_el:
                txt = await precio_el.inner_text()
                data['precio_raw'] = txt
                match = re.search(r'[\d.]+', txt.replace('.', ''))
                if match:
                    data['precio'] = int(match.group())

            features = await page.query_selector_all('.section-icon-features li')
            for f in features:
                txt = await f.inner_text()
                if 'm² tot' in txt.lower():
                    match = re.search(r'(\d+)', txt)
                    if match:
                        data['m2_tot'] = int(match.group(1))
                elif 'm² cub' in txt.lower():
                    match = re.search(r'(\d+)', txt)
                    if match:
                        data['m2_cub'] = int(match.group(1))
                elif 'amb' in txt.lower():
                    match = re.search(r'(\d+)', txt)
                    if match:
                        data['amb'] = int(match.group(1))

            await browser.close()
            return data
    except Exception as e:
        return {'error': str(e)}


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Propiedades']

    updates = []
    zonaprop_urls = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        direccion = row[4] if len(row) > 4 else None
        if not direccion:
            continue

        link = row[36] if len(row) > 36 else None
        if not link:
            continue

        precio = row[6]
        m2c = row[7]

        # Solo scrapear si faltan datos
        if precio and m2c:
            continue

        print(f'Fila {row_num}: {direccion[:40]}...')

        data = None
        if 'argenprop.com' in link:
            data = scrape_argenprop(link)
            time.sleep(1)
        elif 'mercadolibre' in link:
            data = scrape_mercadolibre(link)
            time.sleep(1)
        elif 'zonaprop.com' in link:
            zonaprop_urls.append((row_num, link))
            print(f'  ⏳ Zonaprop (se procesará después)')
            continue
        else:
            print(f'  ⏭️  Dominio no soportado')
            continue

        if data and 'error' in data:
            print(f'  ❌ Error: {data["error"]}')
        elif data:
            print(f'  ✅ {data}')
            updates.append((row_num, data))

    # Procesar Zonaprop (requiere browser)
    if zonaprop_urls:
        print(f'\n=== Procesando {len(zonaprop_urls)} links de Zonaprop ===')

        async def process_zonaprop():
            for row_num, url in zonaprop_urls:
                print(f'Fila {row_num}: {url[:50]}...')
                data = await scrape_zonaprop(url)
                if 'error' in data:
                    print(f'  ❌ Error: {data["error"]}')
                else:
                    print(f'  ✅ {data}')
                    updates.append((row_num, data))

        asyncio.run(process_zonaprop())

    # Actualizar Excel
    print(f'\n=== Actualizando {len(updates)} filas ===')
    for row_num, data in updates:
        if 'precio' in data and not ws.cell(row_num, 7).value:
            ws.cell(row_num, 7).value = data['precio']
            print(f'  Fila {row_num}: precio = {data["precio"]}')
        if 'm2_cub' in data and not ws.cell(row_num, 8).value:
            ws.cell(row_num, 8).value = data['m2_cub']
            print(f'  Fila {row_num}: m2_cub = {data["m2_cub"]}')
        if 'amb' in data and not ws.cell(row_num, 9).value:
            ws.cell(row_num, 9).value = data['amb']

    wb.save(EXCEL_PATH)
    print(f'\n✅ Excel guardado en {EXCEL_PATH}')


if __name__ == '__main__':
    main()
